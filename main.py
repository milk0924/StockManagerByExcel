import openpyxl
import datetime

# 履歴情報からマスタの情報を変更する
def update_by_id(ws, ws_log, ws_person, search_id):
    result, target_row = search_by_id(ws, search_id)
    if result:
        move_quantity, update_datetime, update_person = add_log(ws_log, ws_person, search_id)
                
        # 残数の表示
        current_stock = ws.cell(row=target_row, column=4).value
        print(f"現在の在庫は{current_stock}です")
        # 新しい在庫の表示
        new_stock = current_stock + move_quantity
        # 総入荷数の変動
        if move_quantity>0:
            current_total_quantity = ws.cell(row=target_row, column=6).value
            new_total_quantity = current_total_quantity + move_quantity
            ws.cell(row=target_row, column=6).value = new_total_quantity
            result_comment = f"総入荷数は{new_total_quantity}{ws.cell(row=target_row, column=3).value}です"
        # 総出荷数の変動
        else:
            current_total_shipments = ws.cell(row=target_row, column=7).value
            new_total_shipments = current_total_shipments - move_quantity
            ws.cell(row=target_row, column=7).value = new_total_shipments
            result_comment = f"総出荷数は{new_total_shipments}{ws.cell(row=target_row, column=3).value}です"
        ws.cell(row=target_row, column=4).value = new_stock
        # 最終更新日時の更新
        ws.cell(row=target_row, column=8).value = update_datetime
        # 最終更新人物の更新
        ws.cell(row=target_row, column=9).value = update_person
        print(f"在庫を更新しました．新しい在庫は{new_stock}{ws.cell(row=target_row, column=3).value}です．{result_comment}")

# 履歴シートに履歴を追加する
def add_log(ws_log, ws_person, search_id):
    log_list = []
    target_row = ws_log[1]
    print("履歴の追加を開始します．")
    # IDの入力
    log_list.append(search_id)
    # 変動数の入力 
    move_quantity = int(input(f"{target_row[1].value}を入力してください："))
    log_list.append(move_quantity)
    # 更新日時の入力
    log_list.append(datetime.datetime.now())
    # 更新人物の入力
    person_updated = update_person(ws_person)
    log_list.append(person_updated)
    # メモの入力
    log_memo = input(f"{target_row[4].value}を入力してください")
    log_list.append(log_memo)

    ws_log.append(log_list)
    return log_list[1], log_list[2], log_list[3]

# マスタの編集
def edit_sheet(ws, ws_master_log, ws_person, edit_id):
    result, target_row = search_by_id(ws, edit_id)
    if result:
        while True:
            select_menu = int(input("何を修正しますか？\n1.総入荷数及び総出荷数\t2.総入荷数\t3.総出荷数\t4.最低個数アラート\t5.単位\n0.終了"))
            if select_menu == 1:
                correct_quantity = int(input("総入荷数について，正しい値を入力してください："))
                ws.cell(row=target_row, column=6).value = correct_quantity
                add_master_log(ws_master_log, ws_person, edit_id, ws.cell(row=1, column=6).value, correct_quantity)
                correct_shipments = int(input("総出荷数について，正しい値を入力してください："))
                ws.cell(row=target_row, column=7).value = correct_shipments
                add_master_log(ws_master_log, ws_person, edit_id, ws.cell(row=1, column=7).value, correct_shipments)
                ws.cell(row=target_row, column=4).value = ws.cell(row=target_row, column=6).value - ws.cell(row=target_row, column=7).value
                print(f"修正しました．総入荷数は{ws.cell(row=target_row, column=6).value},総出荷数は{ws.cell(row=target_row, column=7).value},在庫は{ws.cell(row=target_row, column=4).value}になりました．")
            elif select_menu == 2:
                correct_quantity = int(input("総入荷数について，正しい値を入力してください："))
                ws.cell(row=target_row, column=6).value = correct_quantity
                ws.cell(row=target_row, column=4).value = ws.cell(row=target_row, column=6).value - ws.cell(row=target_row, column=7).value
                add_master_log(ws_master_log, ws_person, edit_id, ws.cell(row=1, column=6).value, correct_quantity)
                print(f"修正しました．総入荷数は{ws.cell(row=target_row, column=6).value},在庫は{ws.cell(row=target_row, column=4).value}になりました．")
            elif select_menu == 3:
                correct_shipments = int(input("総出荷数について，正しい値を入力してください："))
                ws.cell(row=target_row, column=7).value = correct_shipments
                ws.cell(row=target_row, column=4).value = ws.cell(row=target_row, column=6).value - ws.cell(row=target_row, column=7).value
                add_master_log(ws_master_log, ws_person, edit_id, ws.cell(row=1, column=7).value, correct_shipments)
                print(f"修正しました．総出荷数は{ws.cell(row=target_row, column=7).value},在庫は{ws.cell(row=target_row, column=4).value}になりました．")
            elif select_menu == 4:
                correct_alert = int(input("最低個数アラートについて，正しい値を入力してください："))
                ws.cell(row=target_row, column=7).value = correct_alert
                add_master_log(ws_master_log, ws_person, edit_id, ws.cell(row=1, column=5).value, correct_alert)
                print(f"修正しました．最低個数アラートは{ws.cell(row=target_row, column=5).value}になりました．")
            elif select_menu == 5:
                correct_unit = str(input("単位について，正しいものを入力してください："))
                ws.cell(row=target_row, column=3).value = correct_unit
                add_master_log(ws_master_log, ws_person, edit_id, ws.cell(row=1, column=3).value, correct_unit)
                print(f"修正しました．単位は{ws.cell(row=target_row, column=3).value}になりました．")
            elif select_menu == 0:
                break
            else: 
                print("そのメニューは存在しません")

# マスタ編集用の履歴管理
def add_master_log(ws_master_log, ws_person, edit_id, edit_title, new_data):
    log_list = []
    print("マスタの修正履歴を登録します")
    # IDの入力
    log_list.append(edit_id)
    # 更新日時
    log_list.append(datetime.datetime.now())
    # 更新人物
    person_update = update_person(ws_person)
    log_list.append(person_update)
    # 更新内容
    edit_content = f"修正内容：{edit_title}を{new_data}に変更"
    log_list.append(edit_content)

    ws_master_log.append(log_list)
    print("登録完了")

# マスタの登録
def add_master(ws, ws_person):
    master_list = []
    print("マスタの登録を開始します")
    # IDの入力（かぶりがあった場合は再度入力を求める）
    while True:
        regist_id = int(input("登録したいIDを入力してください："))
        result, target_row = search_by_id(ws, regist_id)
        if result:
            print(f"ID:{ws.cell(row=target_row, column=1).value}は製品名:{ws.cell(row=target_row, column=2).value}が存在します．")
        else:
            break
    master_list.append(regist_id)
    # 商品名 
    regist_product_name = str(input("商品名を入力してください："))
    master_list.append(regist_product_name)
    # 単位
    regist_unit = str(input("単位を入力してください："))
    master_list.append(regist_unit)
    # 在庫（初期値0）
    master_list.append(int(0))
    # 最低個数アラート
    regist_alert = int(input("最低個数アラートを入力してください："))
    master_list.append(regist_alert)
    # 総入荷数，総出荷数に初期値の0を入れる
    master_list.append(int(0))
    master_list.append(int(0))
    # 最終更新日時
    master_list.append(datetime.datetime.now())
    # 最終更新人物
    person_update = update_person(ws_person)
    master_list.append(person_update)

    ws.append(master_list)
    print("マスタに新たな商品を追加しました．")
    

# パスワードを受け取り，入力者を自動入力する
def update_person(ws_person):
    while True:
        password = int(input("パスワードを入力してください："))
        for row in ws_person.iter_rows(min_row=2, min_col=1, max_col=3):
            if int(row[2].value) == password:
                target_row = row[2].row
                return ws_person.cell(row=target_row, column=1).value
        print("パスワードが不一致です．再度入力してください．")

# IDから製品を検索する
def search_by_id(ws, search_id):
    for row in ws.iter_rows(min_row=2, max_col=1):
        if int(row[0].value) == search_id:
            target_row = row[0].row
            # 製品名の表示
            print(ws.cell(row=target_row, column=2).value)
            return True, target_row
    return False, None

# マスタの列を削除する
def delete_by_id(ws, delete_id):
    result, target_row = search_by_id(ws, delete_id)
    if result:
        ws.delete_rows(target_row)

# データの保存
def save(wb):
    wb.save("stock_manager.xlsx")
    print("データをセーブしました．")

def main():
    wb = openpyxl.load_workbook("stock_manager.xlsx")
    ws = wb["master"]
    ws_log = wb["log"]
    ws_master_log = wb["master_log"]
    ws_person = wb["person"]
    while True:
        select_menu = int(input("行う作業を選んでください\n1.履歴の追加\t2.履歴の修正\t3.マスタの削除\t4.ID検索\t5.マスタの登録\n0.終了\n:"))
        if select_menu == 1:
            search_id = int(input("履歴を追加したい製品IDを入力してください："))
            update_by_id(ws, ws_log, ws_person, search_id)
            save(wb)
        elif select_menu == 2:
            edit_id = int(input("修正したい製品IDを入力してください："))
            edit_sheet(ws, ws_master_log, ws_person, edit_id)
            save(wb)
        elif select_menu == 3:
            delete_id = int(input("削除したい製品IDを入力してください："))
            delete_by_id(ws, delete_id)
            save(wb)
        elif select_menu == 4:
            search_id = int(input("IDを入力してください.そのIDを持つ製品を検索します："))
            search_by_id(ws, search_id)
        elif select_menu == 5:
            add_master(ws, ws_person)
        elif select_menu == 0:
            break
        else:
            print("そのメニューは存在しません")

if __name__ == "__main__":
    main()
